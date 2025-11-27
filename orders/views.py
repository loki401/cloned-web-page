from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.utils import timezone
from .models import Order, OrderItem
from cart.models import Cart
from users.models import Address
import uuid
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

def generate_invoice(order):
    """Generate invoice Excel file for an order"""
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Invoice {order.order_id}"
    
    # Define styles
    title_font = Font(name='Arial', size=20, bold=True, color='2874F0')
    header_font = Font(name='Arial', size=14, bold=True, color='2874F0')
    subheader_font = Font(name='Arial', size=12, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    # Colors
    blue_fill = PatternFill(start_color='2874F0', end_color='2874F0', fill_type='solid')
    light_blue_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
    light_gray_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    # Alignment
    center_align = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    # Border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Company Header
    ws.merge_cells('A1:F1')
    company_cell = ws['A1']
    company_cell.value = "FLIPKART"
    company_cell.font = title_font
    company_cell.alignment = center_align
    
    ws.merge_cells('A2:F2')
    tagline_cell = ws['A2']
    tagline_cell.value = "The Big Billion Day Store"
    tagline_cell.font = Font(name='Arial', size=12, italic=True, color='666666')
    tagline_cell.alignment = center_align
    
    # Invoice Title
    ws.merge_cells('A4:F4')
    invoice_title = ws['A4']
    invoice_title.value = "TAX INVOICE"
    invoice_title.font = header_font
    invoice_title.alignment = center_align
    invoice_title.fill = light_blue_fill
    
    # Invoice Details Section
    ws['A6'] = "Invoice Number:"
    ws['A6'].font = bold_font
    ws['B6'] = order.order_id
    ws['B6'].font = normal_font
    
    ws['D6'] = "Invoice Date:"
    ws['D6'].font = bold_font
    ws['E6'] = order.created_at.strftime('%d-%m-%Y')
    ws['E6'].font = normal_font
    
    ws['A7'] = "Order Date:"
    ws['A7'].font = bold_font
    ws['B7'] = order.created_at.strftime('%d-%m-%Y %H:%M')
    ws['B7'].font = normal_font
    
    ws['D7'] = "Status:"
    ws['D7'].font = bold_font
    ws['E7'] = order.get_status_display()
    ws['E7'].font = normal_font
    
    # Customer Details
    ws['A9'] = "Bill To:"
    ws['A9'].font = subheader_font
    
    customer_name = order.user.get_full_name() or order.user.username
    ws['A10'] = customer_name
    ws['A10'].font = normal_font
    
    ws['A11'] = order.user.email
    ws['A11'].font = normal_font
    
    ws['A12'] = f"Phone: {order.phone_number}"
    ws['A12'].font = normal_font
    
    # Shipping Address
    ws['D9'] = "Ship To:"
    ws['D9'].font = subheader_font
    
    # Split address into lines
    address_lines = order.shipping_address.split('\n')
    for i, line in enumerate(address_lines[:4]):  # Max 4 lines
        if line.strip():
            ws[f'D{10+i}'] = line.strip()
            ws[f'D{10+i}'].font = normal_font
    
    # Items Table Header
    table_start_row = 15
    headers = ['S.No.', 'Product Name', 'Quantity', 'Unit Price', 'Total Price']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=table_start_row, column=col, value=header)
        cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        cell.fill = blue_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # Items Data
    current_row = table_start_row + 1
    subtotal = 0
    
    for i, item in enumerate(order.items.all(), 1):
        # S.No.
        ws.cell(row=current_row, column=1, value=i).font = normal_font
        ws.cell(row=current_row, column=1).alignment = center_align
        ws.cell(row=current_row, column=1).border = thin_border
        
        # Product Name
        ws.cell(row=current_row, column=2, value=item.product.name).font = normal_font
        ws.cell(row=current_row, column=2).alignment = left_align
        ws.cell(row=current_row, column=2).border = thin_border
        
        # Quantity
        ws.cell(row=current_row, column=3, value=item.quantity).font = normal_font
        ws.cell(row=current_row, column=3).alignment = center_align
        ws.cell(row=current_row, column=3).border = thin_border
        
        # Unit Price
        ws.cell(row=current_row, column=4, value=f"₹{item.price}").font = normal_font
        ws.cell(row=current_row, column=4).alignment = right_align
        ws.cell(row=current_row, column=4).border = thin_border
        
        # Total Price
        item_total = item.total_price
        ws.cell(row=current_row, column=5, value=f"₹{item_total}").font = normal_font
        ws.cell(row=current_row, column=5).alignment = right_align
        ws.cell(row=current_row, column=5).border = thin_border
        
        subtotal += item_total
        current_row += 1
        
        # Alternate row coloring
        if i % 2 == 0:
            for col in range(1, 6):
                ws.cell(row=current_row-1, column=col).fill = light_gray_fill
    
    # Summary Section
    summary_start_row = current_row + 1
    
    # Subtotal
    ws.cell(row=summary_start_row, column=4, value="Subtotal:").font = bold_font
    ws.cell(row=summary_start_row, column=4).alignment = right_align
    ws.cell(row=summary_start_row, column=5, value=f"₹{subtotal}").font = bold_font
    ws.cell(row=summary_start_row, column=5).alignment = right_align
    
    # Shipping (assuming free shipping)
    ws.cell(row=summary_start_row + 1, column=4, value="Shipping:").font = normal_font
    ws.cell(row=summary_start_row + 1, column=4).alignment = right_align
    ws.cell(row=summary_start_row + 1, column=5, value="FREE").font = normal_font
    ws.cell(row=summary_start_row + 1, column=5).alignment = right_align
    
    # Total
    ws.cell(row=summary_start_row + 2, column=4, value="Total Amount:").font = Font(name='Arial', size=12, bold=True)
    ws.cell(row=summary_start_row + 2, column=4).alignment = right_align
    ws.cell(row=summary_start_row + 2, column=4).fill = light_blue_fill
    ws.cell(row=summary_start_row + 2, column=5, value=f"₹{order.total_amount}").font = Font(name='Arial', size=12, bold=True)
    ws.cell(row=summary_start_row + 2, column=5).alignment = right_align
    ws.cell(row=summary_start_row + 2, column=5).fill = light_blue_fill
    
    # Footer
    footer_row = summary_start_row + 5
    ws.merge_cells(f'A{footer_row}:E{footer_row}')
    footer_cell = ws[f'A{footer_row}']
    footer_cell.value = "Thank you for shopping with Flipkart!"
    footer_cell.font = Font(name='Arial', size=12, italic=True, color='2874F0')
    footer_cell.alignment = center_align
    
    ws.merge_cells(f'A{footer_row + 1}:E{footer_row + 1}')
    contact_cell = ws[f'A{footer_row + 1}']
    contact_cell.value = "For any queries, contact us at support@flipkart.com | 1800-208-9898"
    contact_cell.font = Font(name='Arial', size=10, color='666666')
    contact_cell.alignment = center_align
    
    # Auto-adjust column widths
    column_widths = [8, 35, 12, 15, 15]  # Predefined widths for better layout
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    return wb

def generate_pdf_invoice(order):
    """Generate PDF invoice for an order"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        textColor=colors.HexColor('#2874f0'),
        alignment=1  # Center alignment
    )
    
    header_style = ParagraphStyle(
        'Header',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=20,
        textColor=colors.HexColor('#2874f0'),
        alignment=1
    )
    
    # Company Header
    title = Paragraph("FLIPKART", title_style)
    elements.append(title)
    
    tagline = Paragraph("The Big Billion Day Store", styles['Normal'])
    tagline.alignment = 1
    elements.append(tagline)
    elements.append(Spacer(1, 20))
    
    # Invoice Title
    invoice_title = Paragraph("TAX INVOICE", header_style)
    elements.append(invoice_title)
    elements.append(Spacer(1, 20))
    
    # Invoice Details Table
    invoice_data = [
        ['Invoice Number:', order.order_id, 'Invoice Date:', order.created_at.strftime('%d-%m-%Y')],
        ['Order Date:', order.created_at.strftime('%d-%m-%Y %H:%M'), 'Status:', order.get_status_display()]
    ]
    
    invoice_table = Table(invoice_data, colWidths=[2*inch, 2*inch, 1.5*inch, 1.5*inch])
    invoice_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    
    elements.append(invoice_table)
    elements.append(Spacer(1, 20))
    
    # Customer Details
    customer_name = order.user.get_full_name() or order.user.username
    
    # Clean shipping address - remove any HTML/template code
    clean_shipping_address = order.shipping_address.replace('<br/>', '\n').replace('<br>', '\n')
    # Remove any remaining HTML tags
    import re
    clean_shipping_address = re.sub(r'<[^>]+>', '', clean_shipping_address)
    
    customer_data = [
        ['Bill To:', 'Ship To:'],
        [f'{customer_name}\n{order.user.email}\nPhone: {order.phone_number}', 
         clean_shipping_address],
    ]
    
    customer_table = Table(customer_data, colWidths=[3.5*inch, 3.5*inch])
    customer_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
    ]))
    
    elements.append(customer_table)
    elements.append(Spacer(1, 30))
    
    # Items Table with Images
    items_data = [['S.No.', 'Product', 'Quantity', 'Unit Price', 'Total Price']]
    
    subtotal = 0
    for i, item in enumerate(order.items.all(), 1):
        # Create product info with image if available
        product_cell_content = []
        
        # Try to add product image
        if item.product.image:
            try:
                from reportlab.platypus import Image
                import os
                
                # Get the full path to the image
                image_path = item.product.image.path
                if os.path.exists(image_path):
                    # Create a small image for the table
                    img = Image(image_path, width=0.6*inch, height=0.6*inch)
                    product_cell_content.append(img)
            except Exception as e:
                pass  # If image fails, just continue without it
        
        # Add product name as paragraph
        product_name = Paragraph(f"<b>{item.product.name}</b>", styles['Normal'])
        product_cell_content.append(product_name)
        
        # If we have both image and text, use a nested table
        if len(product_cell_content) > 1:
            product_info = Table([product_cell_content], colWidths=[0.7*inch, 2.2*inch])
            product_info.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'CENTER'),
                ('ALIGN', (1, 0), (1, 0), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
        else:
            product_info = product_cell_content[0] if product_cell_content else item.product.name
        
        items_data.append([
            str(i),
            product_info,
            str(item.quantity),
            f"Rs.{item.price}",  # Using Rs. instead of ₹ symbol
            f"Rs.{item.total_price}"
        ])
        subtotal += item.total_price
    
    items_table = Table(items_data, colWidths=[0.5*inch, 3*inch, 1*inch, 1.25*inch, 1.25*inch], rowHeights=[None] + [1*inch] * (len(items_data)-1))
    items_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2874f0')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),  # Product column left aligned
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Vertical center alignment
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(items_table)
    elements.append(Spacer(1, 20))
    
    # Summary Table
    summary_data = [
        ['Subtotal:', f"Rs.{subtotal}"],
        ['Shipping:', 'FREE'],
        ['Total Amount:', f"Rs.{order.total_amount}"]
    ]
    
    summary_table = Table(summary_data, colWidths=[5*inch, 2*inch])
    summary_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('FONTNAME', (0, 2), (-1, 2), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 2), (-1, 2), 12),
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('BACKGROUND', (0, 2), (-1, 2), colors.HexColor('#E3F2FD')),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Footer
    footer_text = "Thank you for shopping with Flipkart!"
    footer = Paragraph(footer_text, styles['Normal'])
    footer.alignment = 1
    elements.append(footer)
    
    contact_text = "For any queries, contact us at support@flipkart.com | 1800-208-9898"
    contact = Paragraph(contact_text, styles['Normal'])
    contact.alignment = 1
    elements.append(contact)
    
    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    
    return pdf

@login_required
def checkout(request):
    cart = get_object_or_404(Cart, user=request.user)
    cart_items = cart.items.all()
    
    if not cart_items:
        messages.warning(request, 'Your cart is empty!')
        return redirect('cart:cart')
    
    # Get user's saved addresses
    addresses = Address.objects.filter(user=request.user).order_by('-is_default', '-created_at')
    default_address = addresses.filter(is_default=True).first()
    
    return render(request, 'orders/checkout.html', {
        'cart': cart,
        'cart_items': cart_items,
        'grand_total': cart.total_price,
        'addresses': addresses,
        'default_address': default_address
    })

@login_required
@login_required
def place_order(request):
    if request.method == 'POST':
        from django.db import transaction
        
        try:
            with transaction.atomic():
                cart = get_object_or_404(Cart, user=request.user)
                cart_items = cart.items.select_related('product').all()
                
                if not cart_items:
                    messages.warning(request, 'Your cart is empty!')
                    return redirect('cart:cart')
                
                # Check stock availability first
                for item in cart_items:
                    if item.quantity > item.product.stock:
                        messages.error(request, f'Insufficient stock for {item.product.name}. Only {item.product.stock} left.')
                        return redirect('cart:cart')
                
                # Get shipping address
                address_id = request.POST.get('address_id')
                custom_address = request.POST.get('custom_address', '').strip()
                
                if address_id:
                    # Use saved address
                    try:
                        selected_address = Address.objects.get(id=address_id, user=request.user)
                        shipping_address = f"{selected_address.full_name}\n{selected_address.address_line_1}\n"
                        if selected_address.address_line_2:
                            shipping_address += f"{selected_address.address_line_2}\n"
                        shipping_address += f"{selected_address.city}, {selected_address.state} - {selected_address.pincode}"
                        phone_number = selected_address.phone_number
                    except Address.DoesNotExist:
                        messages.error(request, 'Selected address not found!')
                        return redirect('orders:checkout')
                elif custom_address:
                    # Use custom address
                    shipping_address = custom_address
                    phone_number = request.POST.get('phone', '')
                    
                    # Save address if requested
                    if request.POST.get('save_address'):
                        full_name = request.POST.get('full_name', '')
                        if full_name and phone_number:
                            # Parse address components (simplified)
                            address_parts = custom_address.split('\n')
                            address_line_1 = address_parts[0] if len(address_parts) > 0 else custom_address
                            address_line_2 = address_parts[1] if len(address_parts) > 1 else ''
                            
                            # Create new address
                            Address.objects.create(
                                user=request.user,
                                full_name=full_name,
                                phone_number=phone_number,
                                address_line_1=address_line_1,
                                address_line_2=address_line_2,
                                city='',  # Could be parsed from address
                                state='',  # Could be parsed from address
                                pincode='',  # Could be parsed from address
                                is_default=False
                            )
                else:
                    messages.error(request, 'Please select an address or enter a custom address!')
                    return redirect('orders:checkout')
                
                # Create order
                order = Order.objects.create(
                    user=request.user,
                    order_id=f'FK{uuid.uuid4().hex[:8].upper()}',
                    total_amount=cart.total_price,
                    shipping_address=shipping_address,
                    phone_number=phone_number
                )
                
                # Create order items and deduct stock
                for cart_item in cart_items:
                    # Deduct stock
                    product = cart_item.product
                    product.stock -= cart_item.quantity
                    product.save()
                    
                    OrderItem.objects.create(
                        order=order,
                        product=product,
                        quantity=cart_item.quantity,
                        price=product.discounted_price
                    )
                
                # Clear cart
                cart_items.delete()
                
                # Create order notification
                try:
                    from users.views import create_notification
                    create_notification(
                        user=request.user,
                        title="Order Placed Successfully!",
                        message=f"Your order {order.order_id} has been placed successfully. Total amount: ₹{order.total_amount}",
                        notification_type='order_placed',
                        order_id=order.id,
                        action_url=f'/orders/{order.id}/'
                    )
                except ImportError:
                    pass  # Notification system not available
                
                messages.success(request, f'Order {order.order_id} placed successfully!')
                
                # Store order ID in session for invoice download
                request.session['download_invoice_order_id'] = order.id
                
                return redirect('orders:order_success', order_id=order.id)
                
        except Exception as e:
            messages.error(request, f'An error occurred while placing your order: {str(e)}')
            return redirect('orders:checkout')
    
    return redirect('orders:checkout')

@login_required
def orders(request):
    user_orders = Order.objects.filter(user=request.user).order_by('-created_at')
    return render(request, 'orders/orders.html', {'orders': user_orders})

@login_required
def order_success(request, order_id):
    """Order success page with automatic invoice download"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    return render(request, 'orders/order_success.html', {'order': order})

@login_required
def download_invoice(request, order_id):
    """Download Excel invoice for an order"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    
    # Generate invoice
    wb = generate_invoice(order)
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Invoice_{order.order_id}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response

@login_required
def download_pdf_invoice(request, order_id):
    """Download PDF invoice for an order"""
    order = get_object_or_404(Order, id=order_id, user=request.user)
    
    # Generate PDF invoice
    pdf_data = generate_pdf_invoice(order)
    
    # Create response
    response = HttpResponse(pdf_data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Invoice_{order.order_id}_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    return response

@login_required
def order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    return render(request, 'orders/order_detail.html', {'order': order})
