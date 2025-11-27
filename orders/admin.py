from django.contrib import admin
from django.http import HttpResponse
from django.utils import timezone
from django.urls import path
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from .models import Order, OrderItem


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    extra = 0
    readonly_fields = ['total_price']
    fields = ['product', 'quantity', 'price', 'total_price']

@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    list_display = ['order_id', 'user', 'status', 'total_amount', 'total_items', 'created_at']
    list_filter = ['status', 'created_at', 'updated_at']
    search_fields = ['order_id', 'user__username', 'user__email', 'phone_number']
    readonly_fields = ['order_id', 'created_at', 'updated_at', 'total_items']
    list_editable = ['status']
    inlines = [OrderItemInline]
    actions = ['export_to_csv', 'export_to_excel', 'export_to_pdf', 'mark_as_shipped', 'mark_as_delivered']
    
    fieldsets = (
        ('Order Information', {
            'fields': ('order_id', 'user', 'status', 'total_amount')
        }),
        ('Shipping Details', {
            'fields': ('shipping_address', 'phone_number')
        }),
        ('Timestamps', {
            'fields': ('created_at', 'updated_at'),
            'classes': ('collapse',)
        }),
    )
    
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('export-csv/', self.admin_site.admin_view(self.export_csv_view), name='orders_order_export_csv'),
            path('export-excel/', self.admin_site.admin_view(self.export_excel_view), name='orders_order_export_excel'),
            path('export-pdf/', self.admin_site.admin_view(self.export_pdf_view), name='orders_order_export_pdf'),
        ]
        return custom_urls + urls
    
    def export_to_csv(self, request, queryset):
        """Export selected orders to CSV"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="orders_export_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
        
        writer = csv.writer(response)
        
        # Write header
        writer.writerow([
            'Order ID', 'Customer', 'Email', 'Phone', 'Status', 
            'Total Amount', 'Items Count', 'Order Date', 'Shipping Address'
        ])
        
        # Write data
        for order in queryset:
            writer.writerow([
                order.order_id,
                order.user.get_full_name() or order.user.username,
                order.user.email,
                order.phone_number,
                order.get_status_display(),
                f"Rs.{order.total_amount}",
                order.total_items,
                order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                order.shipping_address.replace('\n', ' ')
            ])
        
        self.message_user(request, f"Successfully exported {queryset.count()} orders to CSV.")
        return response
    
    export_to_csv.short_description = "Export selected orders to CSV"
    
    def export_to_excel(self, request, queryset):
        """Export selected orders to Excel"""
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="orders_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders Report"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2874F0', end_color='2874F0', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal='center', vertical='center')
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = "Orders Export Report"
        title_cell.font = Font(name='Arial', size=16, bold=True, color='2874F0')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add export info
        ws.merge_cells('A2:I2')
        info_cell = ws['A2']
        total_orders = queryset.count()
        total_amount = sum(order.total_amount for order in queryset)
        info_cell.value = f"Total Orders: {total_orders} | Total Amount: Rs.{total_amount:,.2f} | Export Date: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
        info_cell.font = Font(name='Arial', size=10, italic=True)
        info_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Headers
        headers = [
            'Order ID', 'Customer', 'Email', 'Phone', 'Status', 
            'Total Amount', 'Items Count', 'Order Date', 'Shipping Address'
        ]
        
        # Write headers (row 4)
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row, order in enumerate(queryset, 5):
            data = [
                order.order_id,
                order.user.get_full_name() or order.user.username,
                order.user.email,
                order.phone_number,
                order.get_status_display(),
                f"Rs.{order.total_amount}",
                order.total_items,
                order.created_at.strftime('%Y-%m-%d %H:%M:%S'),
                order.shipping_address.replace('\n', ' ')
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = border
                
                # Color alternate rows
                if row % 2 == 0:
                    cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create detailed sheet with order items
        ws2 = wb.create_sheet(title="Order Details")
        
        # Headers for detailed sheet
        detail_headers = [
            'Order ID', 'Customer', 'Product', 'Quantity', 'Unit Price', 
            'Total Price', 'Order Status', 'Order Date'
        ]
        
        # Write headers for detailed sheet
        for col, header in enumerate(detail_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write detailed data
        detail_row = 2
        for order in queryset:
            for item in order.items.all():
                detail_data = [
                    order.order_id,
                    order.user.get_full_name() or order.user.username,
                    item.product.name,
                    item.quantity,
                    f"Rs.{item.price}",
                    f"Rs.{item.total_price}",
                    order.get_status_display(),
                    order.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ]
                
                for col, value in enumerate(detail_data, 1):
                    cell = ws2.cell(row=detail_row, column=col, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    # Color alternate rows
                    if detail_row % 2 == 0:
                        cell.fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
                
                detail_row += 1
        
        # Auto-adjust column widths for detailed sheet
        for col in range(1, len(detail_headers) + 1):
            column_letter = get_column_letter(col)
            max_length = 0
            
            for row in ws2[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        wb.save(response)
        
        self.message_user(request, f"Successfully exported {queryset.count()} orders to Excel with detailed breakdown.")
        return response
    
    export_to_excel.short_description = "Export selected orders to Excel"
    
    def export_to_pdf(self, request, queryset):
        """Export selected orders to PDF"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="orders_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        # Create PDF document
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            textColor=colors.HexColor('#2874f0'),
            alignment=1  # Center alignment
        )
        
        # Title
        title = Paragraph("Orders Export Report", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Summary info
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=10
        )
        
        total_orders = queryset.count()
        total_amount = sum(order.total_amount for order in queryset)
        
        summary_text = f"""
        <b>Export Summary:</b><br/>
        Total Orders: {total_orders}<br/>
        Total Amount: Rs.{total_amount:,.2f}<br/>
        Export Date: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
        """
        
        summary = Paragraph(summary_text, summary_style)
        elements.append(summary)
        elements.append(Spacer(1, 20))
        
        # Orders table
        data = [['Order ID', 'Customer', 'Status', 'Amount', 'Items', 'Date']]
        
        for order in queryset:
            data.append([
                order.order_id,
                order.user.get_full_name() or order.user.username,
                order.get_status_display(),
                f"Rs.{order.total_amount}",
                str(order.total_items),
                order.created_at.strftime('%Y-%m-%d')
            ])
        
        # Create table
        table = Table(data, colWidths=[1.2*inch, 1.5*inch, 1*inch, 1*inch, 0.8*inch, 1*inch])
        
        # Table style
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2874f0')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        
        self.message_user(request, f"Successfully exported {queryset.count()} orders to PDF.")
        return response
    
    export_to_pdf.short_description = "Export selected orders to PDF"
    
    def export_csv_view(self, request):
        """Export all orders to CSV"""
        queryset = Order.objects.all()
        return self.export_to_csv(request, queryset)
    
    def export_excel_view(self, request):
        """Export all orders to Excel"""
        queryset = Order.objects.all()
        return self.export_to_excel(request, queryset)
    
    def export_pdf_view(self, request):
        """Export all orders to PDF"""
        queryset = Order.objects.all()
        return self.export_to_pdf(request, queryset)
    
    def mark_as_shipped(self, request, queryset):
        """Mark selected orders as shipped"""
        updated = queryset.update(status='shipped')
        self.message_user(request, f"Successfully marked {updated} orders as shipped.")
        
        # Create notifications for shipped orders
        try:
            from users.views import create_notification
            for order in queryset:
                create_notification(
                    user=order.user,
                    title="Order Shipped",
                    message=f"Your order {order.order_id} has been shipped and is on its way!",
                    notification_type='order_shipped',
                    order_id=order.id,
                    action_url=f'/orders/{order.id}/'
                )
        except ImportError:
            pass
    
    mark_as_shipped.short_description = "Mark selected orders as shipped"
    
    def mark_as_delivered(self, request, queryset):
        """Mark selected orders as delivered"""
        updated = queryset.update(status='delivered')
        self.message_user(request, f"Successfully marked {updated} orders as delivered.")
        
        # Create notifications for delivered orders
        try:
            from users.views import create_notification
            for order in queryset:
                create_notification(
                    user=order.user,
                    title="Order Delivered",
                    message=f"Your order {order.order_id} has been delivered successfully!",
                    notification_type='order_delivered',
                    order_id=order.id,
                    action_url=f'/orders/{order.id}/'
                )
        except ImportError:
            pass
    
    mark_as_delivered.short_description = "Mark selected orders as delivered"


@admin.register(OrderItem)
class OrderItemAdmin(admin.ModelAdmin):
    list_display = ['order', 'product', 'quantity', 'price', 'total_price']
    list_filter = ['order__status', 'order__created_at']
    search_fields = ['order__order_id', 'product__name']
    readonly_fields = ['total_price']
